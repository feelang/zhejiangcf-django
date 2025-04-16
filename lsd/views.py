import logging
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from .models import LsdSurvey, LsdOrganization
from django.core.serializers import serialize
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import openpyxl
from django.contrib import messages
from django.db.models import Q, Count
from .models import UserProfile
import re
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from . import constants

logger = logging.getLogger('log')

@login_required
def index(request):
    return render(request, 'lsd/index.html')

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def create_survey(request):
    try:
        data = json.loads(request.body)
        
        # 获取用户所属机构
        try:
            user_profile = request.user.profile
            organization = user_profile.organization if user_profile and user_profile.organization else None
            organization_code = organization.code if organization else None
        except UserProfile.DoesNotExist:
            return JsonResponse({
                'code': 403,
                'message': constants.ERROR_USER_NO_ORGANIZATION
            }, status=403)

        # 创建新的调查记录
        survey = LsdSurvey(
            _openId=f"web_{request.user.id}",
            name=data.get('name'),
            age=data.get('age'),
            phone=data.get('phone'),
            organization=organization_code,
            occupation=data.get('occupation'),
            project=data.get('project'),
            groupSelection=data.get('groupSelection'),
            sexualExperience=data.get('sexualExperience'),
            cervicalCancerScreening=data.get('cervicalCancerScreening'),
            hpv_result=data.get('hpv_result'),
            tct_result=data.get('tct_result'),
            biopsy_result=data.get('biopsy_result'),
            remark=data.get('remark')
        )

        # 保存到数据库
        survey.save()

        return JsonResponse({
            'code': 0,
            'message': constants.SUCCESS_CREATE,
            'data': {
                'id': survey.id
            }
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'code': 400,
            'message': constants.ERROR_INVALID_JSON
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'code': 500,
            'message': str(e)
        }, status=500)

@require_http_methods(["GET"])
def list_surveys_weapp(request):
    try:
        # 从请求头获取微信云托管注入的openid
        openId = request.headers.get('X-WX-OPENID') or request.headers.get('X-WX-FROM-OPENID')
        logger.info(f'openId: {openId}')  # Add this line for debuggin
        if not openId:
            return JsonResponse({
                'code': 401,
                'message': '未授权访问'
            }, status=401)

        # 查询该用户的所有问卷
        surveys = LsdSurvey.objects.filter(_openId=openId).values(
            'id', '_openId', 'name', 'age', 'phone', 'organization',
            'occupation', 'project', 'groupSelection',
            'sexualExperience', 'cervicalCancerScreening',
            'created_at', 'updated_at'
        )

        return JsonResponse({
            'code': 0,
            'message': '查询成功',
            'data': list(surveys)
        })
    except Exception as e:
        return JsonResponse({
            'code': 500,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def update_survey(request, survey_id):
    try:
        data = json.loads(request.body)
        survey = LsdSurvey.objects.get(id=survey_id)
        
        # 更新问卷数据
        if data.get('name') is not None:
            survey.name = data.get('name')
        if data.get('age') is not None:
            survey.age = data.get('age')
        if data.get('phone') is not None:
            survey.phone = data.get('phone')
        if data.get('organization') is not None:
            survey.organization = data.get('organization')
        if data.get('occupation') is not None:
            survey.occupation = data.get('occupation')
        if data.get('project') is not None:
            survey.project = data.get('project')
        if data.get('groupSelection') is not None:
            survey.groupSelection = data.get('groupSelection')
        if data.get('sexualExperience') is not None:
            survey.sexualExperience = data.get('sexualExperience')
        if data.get('cervicalCancerScreening') is not None:
            survey.cervicalCancerScreening = data.get('cervicalCancerScreening')
        if data.get('hpv_result') is not None:
            survey.hpv_result = data.get('hpv_result')
        if data.get('tct_result') is not None:
            survey.tct_result = data.get('tct_result')
        if data.get('biopsy_result') is not None:
            survey.biopsy_result = data.get('biopsy_result')
        if data.get('remark') is not None:
            survey.remark = data.get('remark')
        survey.save()
        
        return JsonResponse({
            'code': 0,
            'message': '更新成功'
        })
    except LsdSurvey.DoesNotExist:
        return JsonResponse({
            'code': 404,
            'message': '问卷不存在'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'code': 500,
            'message': str(e)
        }, status=500)

def is_survey_complete(survey):
    """检查调查数据是否完整"""
    return all([
        survey.name,
        survey.phone and re.match(r'^1[3-9]\d{9}$', survey.phone),
        survey.age and survey.age >= 25,
        survey.organization,
        survey.occupation,
        survey.project,
        survey.groupSelection,
        survey.sexualExperience is not None,
        survey.cervicalCancerScreening is not None,
        survey.hpv_result,
        survey.tct_result,
        survey.biopsy_result
    ])

@login_required
def survey_list(request):
    # 检查用户是否属于lsd组
    is_lsd_user = request.user.groups.filter(name='lsd').exists()
    if not is_lsd_user:
        messages.error(request, constants.ERROR_NO_PERMISSION)
        return redirect('lsd:index')

    is_admin = request.user.is_staff or request.user.is_superuser
    if not is_admin:
        # 获取普通用户所属机构
        try:
            user_profile = request.user.profile
            organization = user_profile.organization if user_profile and user_profile.organization else None
        except UserProfile.DoesNotExist:
            organization = None

        if not organization:
            messages.error(request, constants.ERROR_NO_ORGANIZATION)
            return redirect('lsd:index')
    else:
        organization = None

    # 获取搜索参数
    search_query = request.GET.get('q', '')

    # 获取数据列表
    if is_admin:
        # lsd组用户可以查看所有问卷
        surveys = LsdSurvey.objects.all()
    else:
        # 普通用户只能查看自己机构的问卷
        surveys = LsdSurvey.objects.filter(organization=organization.code)

    # 如果有搜索参数，添加过滤条件
    if search_query:
        surveys = surveys.filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(hpv_result__icontains=search_query) |
            Q(tct_result__icontains=search_query)
        )

    # 按更新时间倒序排序
    surveys = surveys.order_by('-updated_at')

    # 分页
    paginator = Paginator(surveys, 10)  # 每页显示10条记录
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 为每个调查添加完整性标记
    for survey in page_obj:
        survey.is_complete = is_survey_complete(survey)

    return render(request, 'lsd/survey_list.html', {
        'page_obj': page_obj,
        'is_paginated': paginator.num_pages > 1,
        'organization': organization,
        'is_lsd_user': is_lsd_user
    })

@login_required
def import_surveys(request):
    # 获取用户所属机构
    try:
        user_profile = request.user.profile
        organization = user_profile.organization if user_profile and user_profile.organization else None
    except UserProfile.DoesNotExist:
        organization = None

    if not organization:
        messages.error(request, constants.ERROR_NO_ORGANIZATION)
        return redirect('lsd:index')

    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, constants.ERROR_NO_EXCEL_FILE)
            return redirect('lsd:import_surveys')

        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, constants.ERROR_INVALID_EXCEL_FORMAT)
            return redirect('lsd:import_surveys')

        try:
            # 读取Excel文件
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # 验证必要的列是否存在
            headers = [str(cell.value).strip() for cell in sheet[1]]
            missing_columns = [col for col in constants.REQUIRED_COLUMNS if col not in headers]
            if missing_columns:
                messages.error(request, constants.ERROR_MISSING_COLUMNS.format(columns=", ".join(missing_columns)))
                return redirect('lsd:import_surveys')

            # 获取列索引
            column_indices = {col: headers.index(col) + 1 for col in constants.REQUIRED_COLUMNS}

            # 导入数据
            success_count = 0
            error_count = 0
            skipped_count = 0
            invalid_phone_count = 0
            invalid_phone_records = []

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # 从第二行开始（跳过表头）
                try:
                    # 获取行数据
                    row_data = {col: str(row[idx-1].value).strip() if row[idx-1].value is not None else '' for col, idx in column_indices.items()}

                    name = str(row_data['姓名']).strip()
                    phone = str(row_data['电话']).strip()

                    # 检查姓名和电话是否为空
                    if not name and not phone:
                        skipped_count += 1
                        logger.info(f'跳过空数据行: 行号 {row_idx}, 姓名: {row_data["姓名"]}, 电话: {row_data["电话"]}')
                        continue

                    if not name:
                        name = ''
                    
                    # 处理并验证手机号
                    phone = str(phone).replace('.0', '')  # 处理Excel可能将数字加上.0的情况
                    is_valid_phone = bool(re.match(r'^1[3-9]\d{9}$', phone))  # 验证中国大陆手机号格式

                    if not is_valid_phone:
                        invalid_phone_count += 1
                        invalid_phone_records.append(f"行号 {row_idx}: {row_data['姓名']} ({phone})")
                        logger.info(f'无效手机号: 行号 {row_idx}, 姓名: {row_data["姓名"]}, 电话: {phone}')

                    # 转换布尔值
                    sexual_experience = str(row_data['是否有过性生活']).lower() in ['是', '有', 'yes', 'true', '1']
                    cervical_screening = str(row_data['一年内是否做过宫颈癌筛查']).lower() in ['是', '有', 'yes', 'true', '1']

                    # 转换年龄为整数
                    try:
                        age = int(float(row_data['年龄']))
                    except (ValueError, TypeError):
                        age = 0

                    # 创建或更新调查记录
                    survey, created = LsdSurvey.objects.update_or_create(
                        phone=phone,
                        name=row_data['姓名'],
                        defaults={
                            '_openId': f"import_{phone}",
                            'age': age,
                            'organization': organization.code,
                            'occupation': row_data['职业'],
                            'project': constants.PROJECT_NAME,
                            'groupSelection': row_data['群体选择'],
                            'sexualExperience': sexual_experience,
                            'cervicalCancerScreening': cervical_screening,
                            'hpv_result': row_data['本次活动-HPV结果'],
                            'tct_result': row_data['本次活动-TCT结果'],
                            'biopsy_result': row_data['活检结果'],
                            'remark': row_data['备注']
                        }
                    )

                    success_count += 1
                except Exception as e:
                    error_count += 1
                    logger.error(f'导入数据失败: {str(e)}, 行号: {row_idx}, 数据: {row_data}')

            # 构建消息
            message = f'成功导入 {success_count} 条数据，失败 {error_count} 条，跳过 {skipped_count} 条空数据'
            if invalid_phone_count > 0:
                message += f'，其中 {invalid_phone_count} 条手机号格式不正确'
                if len(invalid_phone_records) <= 5:  # 如果无效手机号记录不多，直接显示
                    message += f'：{", ".join(invalid_phone_records)}'
                else:  # 如果无效手机号记录较多，只显示前5条
                    message += f'：{", ".join(invalid_phone_records[:5])}等'

            messages.success(request, message)
            return redirect('lsd:import_surveys')

        except Exception as e:
            messages.error(request, f'导入失败: {str(e)}')
            return redirect('lsd:import_surveys')

    return render(request, 'lsd/import_surveys.html', {
        'organization': organization,
    })

@login_required
@require_http_methods(["POST"])
def delete_survey(request, survey_id):
    try:
        survey = LsdSurvey.objects.get(id=survey_id)
        
        # 检查用户权限
        if not request.user.is_staff:
            # 获取用户所属机构
            try:
                user_profile = request.user.profile
                organization = user_profile.organization if user_profile and user_profile.organization else None
                organization_code = organization.code if organization else None
                
                # 如果用户不是staff且问卷不属于用户所在机构，则拒绝删除
                if organization_code != survey.organization:
                    return JsonResponse({
                        'code': 403,
                        'message': constants.ERROR_NO_PERMISSION_DELETE
                    }, status=403)
            except UserProfile.DoesNotExist:
                return JsonResponse({
                    'code': 403,
                    'message': constants.ERROR_NO_PERMISSION_DELETE
                }, status=403)
        
        # 删除问卷
        survey.delete()
        
        return JsonResponse({
            'code': 0,
            'message': constants.SUCCESS_DELETE
        })
    except LsdSurvey.DoesNotExist:
        return JsonResponse({
            'code': 404,
            'message': constants.ERROR_SURVEY_NOT_FOUND
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'code': 500,
            'message': str(e)
        }, status=500)

@login_required
def create_survey_page(request):
    # 获取用户所属机构
    try:
        user_profile = request.user.profile
        organization = user_profile.organization if user_profile and user_profile.organization else None
    except UserProfile.DoesNotExist:
        organization = None

    if not organization:
        messages.error(request, '您未关联机构信息，请联系管理员进行关联')
        return redirect('lsd:index')

    return render(request, 'lsd/create_survey.html', {
        'organization': organization
    })

@login_required
def statistics(request):
    # 检查用户是否是管理员
    if not request.user.is_staff:
        messages.error(request, '您没有权限访问此页面')
        return redirect('lsd:index')

    # 获取问卷总数
    total_surveys = LsdSurvey.objects.count()

    # 按机构统计问卷数量
    surveys_by_org = LsdSurvey.objects.values('organization').annotate(count=Count('id')).order_by('-count')

    # 获取机构名称
    org_names = {}
    for org in LsdOrganization.objects.all():
        org_names[org.code] = org.name

    # 添加机构名称到统计结果中
    for survey in surveys_by_org:
        survey['org_name'] = org_names.get(survey['organization'], '未知机构')

    context = {
        'total_surveys': total_surveys,
        'surveys_by_org': surveys_by_org,
    }
    return render(request, 'lsd/statistics.html', context)

@login_required
def export_surveys(request):
    # 检查用户是否是管理员
    if not request.user.is_staff:
        messages.error(request, '您没有权限执行此操作')
        return redirect('lsd:index')

    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "问卷数据"

    # 定义表头
    headers = [
        '姓名', '年龄', '手机号', '机构', '职业', '项目', '分组选择',
        '性经历', '宫颈癌筛查', 'HPV结果', 'TCT结果', '活检结果', '备注',
        '创建时间', '更新时间'
    ]

    # 设置表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 获取所有问卷数据
    surveys = LsdSurvey.objects.all().order_by('-created_at')

    # 写入数据
    for row_num, survey in enumerate(surveys, 2):
        ws.cell(row=row_num, column=1, value=survey.name)
        ws.cell(row=row_num, column=2, value=survey.age)
        ws.cell(row=row_num, column=3, value=survey.phone)
        ws.cell(row=row_num, column=4, value=survey.organization)
        ws.cell(row=row_num, column=5, value=survey.occupation)
        ws.cell(row=row_num, column=6, value=survey.project)
        ws.cell(row=row_num, column=7, value=survey.groupSelection)
        ws.cell(row=row_num, column=8, value='是' if survey.sexualExperience else '否')
        ws.cell(row=row_num, column=9, value='是' if survey.cervicalCancerScreening else '否')
        ws.cell(row=row_num, column=10, value=survey.hpv_result or '')
        ws.cell(row=row_num, column=11, value=survey.tct_result or '')
        ws.cell(row=row_num, column=12, value=survey.biopsy_result or '')
        ws.cell(row=row_num, column=13, value=survey.remark or '')
        ws.cell(row=row_num, column=14, value=survey.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=15, value=survey.updated_at.strftime('%Y-%m-%d %H:%M:%S'))

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 设置所有单元格居中对齐
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 创建响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'问卷数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # 保存工作簿到响应
    wb.save(response)
    return response

@login_required
def edit_survey_page(request, survey_id):
    try:
        survey = LsdSurvey.objects.get(id=survey_id)
        return render(request, 'lsd/edit_survey.html', {
            'survey': survey
        })
    except LsdSurvey.DoesNotExist:
        messages.error(request, '问卷不存在')
        return redirect('lsd:survey_list')
